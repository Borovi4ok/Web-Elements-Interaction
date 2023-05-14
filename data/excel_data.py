import openpyxl
import pytest


# extract data from an Excel spreadsheet based on test_case_name, for data driven tests
def get_excel_row_data(test_case_name):
    # create dictionary to extract data from excel
    excel_data_dict = {}

    # create var and assign to it path to the Excel workbook with openpyxl.load_workbook() method
    book = openpyxl.load_workbook("C:\\Users\\bogod\\PycharmProjects\\pythonProject\\WebInteractionDemoQA\\data\\test_data_excel.xlsx")

    # work with active sheet in the Excel document
    sheet = book.active

    # extract all values without titles
    # in range(2, ...) - range starts from 2 because we don't need the title of the row or column but values only
    for i in range(2, sheet.max_row + 1):  # to get rows (runs from up to down)
        if sheet.cell(row=i, column=1).value == test_case_name:
            for j in range(2, sheet.max_column + 1):  # to get columns (left to right)

                # create a dictionary of key-value pairs
                excel_data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    # print(excel_data_dict)
    return excel_data_dict


# create fixture with wrapper for get_excel_row_data() to get a particular value for 'key' from returned dictionary
@pytest.fixture
def get_excel_data():
    def wrapper(test_case_name, key):
        test_case_dict = get_excel_row_data(test_case_name)
        return test_case_dict[key]
    return wrapper
